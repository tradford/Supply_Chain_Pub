import pandas as pd
import os
from dateutil.relativedelta import relativedelta
from logMessage import logMessage as lm
from sqlalchemy import create_engine
import shutil
import urllib
import os
import pyodbc
import requests
from msTeam import msteam
import datetime
from datetime import datetime as dt
import openpyxl


''' TODO: Pull data from excel file and load into database'''
         
#df = pd.read_excel(r'path"

#####SET VARIABLES
today = dt.today()
tody = today.strftime('%b%y')
if tody[0:3] == 'Oct' or tody[0:3] == 'Nov' or tody[0:3] == 'Dec':
    yr = int(tody[3:]) + 1
    sheetnm = '20' + tody[3:] + '-' + str(yr) 
else:
    yrs = int(tody[3:]) -1
    sheetnm = '20' +  str(yrs) + '-' + tody[3:]

print(sheetnm)
LOG = r"path"
site = "path"

#####GET TOKEN FOR SHAREPOINT

try:
    url_toke = "path"

    payload='grant_type='
    headers = {
    'Content-Type': 'application/x-www-form-urlencoded',
    'Cookie': ''
    }

    response = requests.request("GET", url_toke, headers=headers, data=payload)

    #print(response.text)
    token = response.json()

    ###### Step 3 Access site and copy Formula data to a local folder

    access = "Bearer " + token["access_token"]

    #url = "path"

    url = "path"

    payload={}
    headers = {
    'Authorization': access
    }

    response = requests.request("GET", url, headers=headers, data=payload)
    response_j = response.json()
    print(response.text)

    def find_id(data, target_id):
        for key, value in data.items():
            if key == 'id' and value == target_id:
                return data
            elif isinstance(value, dict):
                found = find_id(value, target_id)
                if found:
                    return found
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        found = find_id(item, target_id)
                        if found:
                            return found

    target_id = ''
    result = find_id(response_j, target_id)
    # if result:
    #     print(f"Found ID '{target_id}' in the JSON data: {result}")
    # else:
    #     print(f"ID '{target_id}' not found in the JSON data.")

except Exception as e:
    msteam(site, "Product Formula file is not found for Supply Chain Automation, check the SharePoint ID via MSGRAPH API the code is looking for 01G6TXF6XHBRJYNYVGYRBK3UU2F5W6X7XM, see file update_esi_product_weight.py: Error: " + str(e),"Trent Radford")
    lm(LOG, "Product Formula file not found for Supply Chain Automation Error, check the SharePoint ID via MSGRAPH API the code is looking for 01G6TXF6XHBRJYNYVGYRBK3UU2F5W6X7XM : " + str(e))
    


downld = response.json()

filename = ""
try:
    downld_url = result['@microsoft.graph.downloadUrl']


    cwd = r'path"
    response = requests.get(downld_url)
    os.chdir(cwd)
    filename = ""
    open(filename, "wb").write(response.content)

except Exception as e:
    msteam(site, "Product Formula file has been found in SharePoint but cannot be written to C__Users__esisvc__Projects__OPS__Supply_Chain__data see file update_esi_product_weight.py: Error: " + str(e),"Trent Radford")
    lm(LOG, "Product Formula file cannot be written for Supply Chain Automation Error: " + str(e))
    
    
    


def read_comments(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in workbook.sheetnames:  #202408098
        sheet = workbook[sheet_name]
        print(f"Reading comments from worksheet '{sheet_name}':")
        
        # Loop through all rows and columns
        for row in sheet.iter_rows():
            for cell in row:
                # Check if the cell has a comment
                if cell.comment is not None:
                    comment = cell.comment
                    print(f"Cell {cell.coordinate}: Comment by {comment.author}: {comment.text}")

file_path = fr"path"
read_comments(file_path)

def load_excel_with_comments(file_path, sheet_name):
    # Load workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]

    # Load the Excel sheet into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    #df.columns = df.iloc[1]
    # Initialize a new column for comments
    df['Comment'] = None

    # Loop through all rows and columns
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell has a comment
            if cell.comment is not None:
                comment = cell.comment
                
                comment_text = f"Cell {cell.coordinate}: {comment.text}"
                
                # Get the index of the row where the comment is made
                row_index = cell.row - 1
                # Add the comment to the 'Comment' column in the DataFrame
                df.at[row_index, 'Comment'] = comment_text

    return df

#file_path = "your_file_path.xlsx"
# pd.set_option('display.max_columns', None)
# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_colwidth', None)
sheet_name = "Sheet1"  # Change to the name of your sheet
df = load_excel_with_comments(file_path, sheet_name)



df['Comment'] = df['Comment'].str.extract(r'(Values verified.*)')
df.loc[0, 'Comment'] = "Comment"
# df.columns = df.iloc[1]
# df.drop(index=df.index[:1],inplace=True)
# print(df.head(20))
# print(df.columns.values.tolist())
new_header = df.iloc[0] #grab the first row for the header
df = df[1:] #take the data less the header row
df.columns = new_header
df.dropna(axis=1,how='all', inplace=True)
df.dropna(axis=0,how='all', inplace=True)
df.rename(columns={"Weight\nlbs/gal.": 'Pounds_Per_Gal', 'gal./ton': 'Gal_Per_Ton' }, inplace=True)
df.fillna({'Part Number' :"", 'Master Part Description' : "", 'Inventory Part Description':"", 'Sales Part Description': "", 'SG': 0 , 'Pounds_Per_Gal':0, 'Gal_Per_Ton' :0 ,'Comment' : ""} , inplace=True)
print(df)
print(df.columns.values.tolist())
# df['Revision Date'] = df['Revision Date'].astype(str).replace('NaN', 'null')
def format_date(date_input):
    if isinstance(date_input, pd.Timestamp):
        return date_input.strftime('%d-%b-%Y')
    else:
        return date_input
 
df['Revision Date'] = df['Revision Date'].apply(format_date)
# ######## Establish a connection to the database

# lib_dir = r"path"
# os.chdir(lib_dir)
# #try:
# #engine = create_engine('oracle+cx_oracle://{username}:{password}@{hostname}:{port}/{database}')



conn_str = "DRIVER={SQL Server};SERVER=;DATABASE="
quoted_conn_str = urllib.parse.quote_plus(conn_str)
engine = create_engine(f"mssql+pyodbc:///?odbc_connect={quoted_conn_str}")

df.to_sql('product_weights', con=engine,  if_exists='replace', index=False)