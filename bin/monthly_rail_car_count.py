import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

# Read the Excel file
file_path = r'path"
df = pd.read_excel(file_path, sheet_name='Leased Fleet')

# Filter rows with active status
active_rows = df[df["Status"] == "Active"]

# Count active statuses for each Car Type
active_count = active_rows["Car Type"].value_counts().reset_index()

# Rename columns
active_count.columns = ["Car Type", "active_count"]

# Add the current month column
current_month = datetime.today()
prev_month = (current_month - relativedelta(months=1)).strftime('%B %Y')
active_count["Month"] = prev_month

# Read the existing data from the "Active_Counts" sheet
existing_data = pd.read_excel(file_path, sheet_name="Active_Counts")

# Concatenate the new data with the existing data
combined_data = pd.concat([existing_data, active_count], ignore_index=True)

combined_data['Month'] = pd.to_datetime(combined_data['Month']).dt.strftime('%d-%b-%y')
# Remove duplicates based on the "Car Type" and "Month" columns
updated_data = combined_data.drop_duplicates(subset=["Car Type", "Month"])

# Check if the "Active_Counts" sheet exists and remove it
workbook = load_workbook(file_path)
if "Active_Counts" in workbook.sheetnames:
    sheet = workbook["Active_Counts"]
    workbook.remove(sheet)
    workbook.save(file_path)

# Save the updated DataFrame to the "Active_Counts" sheet in the Excel file
with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
    updated_data.to_excel(writer, sheet_name="Active_Counts", index=False)

