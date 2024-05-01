import pandas as pd
from datetime import datetime
import openpyxl

# Define the Excel file path
file_path = r'path"

# Load the 'Leased Fleet' sheet into a DataFrame
df = pd.read_excel(file_path, sheet_name='Leased Fleet')

# Check if today is the last day of the month
today = datetime.today()
# if today.month != (today.replace(day=28) + pd.DateOffset(days=4)).day.month:

# Count the number of 'Tank Car' and 'Hopper' with 'Active' status
tank_car_count = df[(df['Car Type'] == 'Tank Car') & (df['Status'] == 'Active')].shape[0]
hopper_count = df[(df['Car Type'] == 'Hopper') & (df['Status'] == 'Active')].shape[0]

# Prepare data to append
data_to_append = {
    'Car Type': ['Tank Car', 'Hopper'],
    'active_count': [tank_car_count, hopper_count],
    'Month': [today.strftime('%m/%d/%Y') for _ in range(2)]
}
append_df = pd.DataFrame(data_to_append)

# Append data to 'Active_Counts' sheet
book = openpyxl.load_workbook(file_path)



if 'Active_Counts' not in book.sheetnames:
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    append_df.to_excel(writer, index=False, sheet_name='Active_Counts')
else:
    # Append data to 'Active_Counts' sheet without writing headers
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        writer.book = book
        # Get the last row in the existing 'Active_Counts' sheet
        startrow = book['Active_Counts'].max_row
        # Write the new data
        append_df.to_excel(writer, index=False, sheet_name='Active_Counts', startrow=startrow, header=False)

# Save and close the workbook


print("Data appended successfully.")
# else:
#     print("Today is not the last day of the month. No action taken.")
